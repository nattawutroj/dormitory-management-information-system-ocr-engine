from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from io import BytesIO
from PIL import Image
from ultralytics import YOLO
import cv2
import numpy as np
import easyocr
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import os
import shutil
import qrcode
from PIL import Image, ImageDraw
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)
CORS(app)  # Allow CORS for localhost

# Load YOLO model and OCR reader
model = YOLO("./src/model/best.pt")
reader = easyocr.Reader(['en'])  # Using English language for OCR

def generate_dummy_data(num_rows):
    data_list = {"row": {}}
    for i in range(1, num_rows + 1):
        data_list["row"][str(i)] = {
            "column": {
                "A": str(i),
                "B": "2000", #เล่มที่
                "C": "10", #เลขที่
                "D": "- ค่าไฟฟ้าเดือน มีนาคม", #รายการ
                "E": "3", #เลขตึก
                "F": "104", #เลขห้อง
                "G": f"นายทดสอบ {i}", #ชื่อ
                "H": "30", #จำนวนยูนิต
                "I": "6", #ราคาต่อยูนิต
                "J": "180", #ยอดราคาใช้งาน
                "L": "2", #จำนวนคนในห้องพัก
                "M": "90", #จำนวนเงินที่หารคน
                "N": "100", #จำนวนเงินค่าปรับ
                "O": "190" #จำนวนเงินรวม
            }
        }
    return data_list

def format_data_list(data_list):
    """Convert list of data to the required format"""
    formatted_data = {"row": {}}
    for i, row_data in enumerate(data_list, 1):
        formatted_data["row"][str(i)] = {
            "column": {
                "A": str(i),
                "B": str(row_data.get("book_number", "")),  # เล่มที่
                "C": str(row_data.get("number", "")),  # เลขที่
                "D": str(row_data.get("description", "")),  # รายการ
                "E": str(row_data.get("building_number", "")),  # เลขตึก
                "F": str(row_data.get("room_number", "")),  # เลขห้อง
                "G": str(row_data.get("name", "")),  # ชื่อ
                "H": str(row_data.get("units", "")),  # จำนวนยูนิต
                "I": str(row_data.get("unit_price", "")),  # ราคาต่อยูนิต
                "J": str(row_data.get("usage_price", "")),  # ยอดราคาใช้งาน
                "L": str(row_data.get("people_count", "")),  # จำนวนคนในห้องพัก
                "M": str(row_data.get("price_per_person", "")),  # จำนวนเงินที่หารคน
                "N": str(row_data.get("penalty", "")),  # จำนวนเงินค่าปรับ
                "O": str(row_data.get("total", ""))  # จำนวนเงินรวม
            }
        }
    return formatted_data

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
        
        target_cell.number_format = source_cell.number_format

def copy_border(source_border):
    """Create a new border object from source border"""
    if not source_border:
        return None
    
    return Border(
        left=Side(style=source_border.left.style) if source_border.left else None,
        right=Side(style=source_border.right.style) if source_border.right else None,
        top=Side(style=source_border.top.style) if source_border.top else None,
        bottom=Side(style=source_border.bottom.style) if source_border.bottom else None
    )

def parse_range(range_string):
    """Parse a range string into min_col, min_row, max_col, max_row"""
    start, end = range_string.split(':')
    start_col, start_row = coordinate_from_string(start)
    end_col, end_row = coordinate_from_string(end)
    
    return (
        column_index_from_string(start_col),
        start_row,
        column_index_from_string(end_col),
        end_row
    )

def setup_sheet_template(sheet, template_sheet, start_row=1):
    """Copy template settings to new sheet"""
    # Copy print settings
    sheet.page_setup = template_sheet.page_setup
    sheet.print_options = template_sheet.print_options
    sheet.print_title_rows = template_sheet.print_title_rows
    sheet.print_title_cols = template_sheet.print_title_cols
    sheet.print_area = template_sheet.print_area

    # Store merged cells ranges in a list
    merged_ranges = []
    for merged_cells in template_sheet.merged_cells.ranges:
        min_row = merged_cells.min_row + start_row - 1
        max_row = merged_cells.max_row + start_row - 1
        min_col = merged_cells.min_col
        max_col = merged_cells.max_col
        new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        merged_ranges.append(new_range)

    # Apply merged cells after collecting all ranges
    for new_range in merged_ranges:
        sheet.merge_cells(new_range)

    # Copy all cells from template with offset
    for row in template_sheet.rows:
        for cell in row:
            if not isinstance(cell, MergedCell):
                new_row = cell.row + start_row - 1
                col_letter = get_column_letter(cell.column)
                target_cell = sheet[f"{col_letter}{new_row}"]
                target_cell.value = cell.value
                copy_cell_style(cell, target_cell)

    # Copy column dimensions
    for col in list(template_sheet.column_dimensions.keys()):
        sheet.column_dimensions[col] = template_sheet.column_dimensions[col]

    # Copy row dimensions with offset
    for row in list(template_sheet.row_dimensions.keys()):
        new_row = row + start_row - 1
        sheet.row_dimensions[new_row] = template_sheet.row_dimensions[row]

    # Create thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply borders to all cells in the section
    for row in range(start_row, start_row + 25):
        for col in range(1, 21):  # A to T
            cell = sheet[f"{get_column_letter(col)}{row}"]
            # Copy border from template if it exists
            template_row = row - start_row + 1
            template_cell = template_sheet[f"{get_column_letter(col)}{template_row}"]
            if template_cell.border:
                cell.border = copy_border(template_cell.border)
            else:
                cell.border = thin_border

    # Ensure borders are applied to merged cells
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = parse_range(merged_range)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet[f"{get_column_letter(col)}{row}"]
                if not cell.border:
                    cell.border = thin_border

def fill_room_names(sheet, rooms, start_index, start_row=1):
    """Fill room names in the specified positions on the sheet"""
    positions = ['A5', 'A9', 'A13', 'A17']
    for i, pos in enumerate(positions):
        if start_index + i < len(rooms):
            # Calculate new position with offset
            col = pos[0]
            row = int(pos[1:]) + start_row - 1
            cell = sheet[f"{col}{row}"]
            cell.value = rooms[start_index + i]['room_name']
            cell.font = Font(name='TH SarabunPSK', size=14)

def fill_student_data(sheet, room, start_row):
    """Fill student data for a room group starting at the specified row"""
    students = room['student_list']
    
    # Fill student data for up to 4 students
    for i, student in enumerate(students[:4]):
        row = start_row + i
        
        # Student ID
        sheet[f'B{row}'] = student['student_id']
        
        # Name
        sheet[f'C{row}'] = student['name']
        
        # Initial
        sheet[f'D{row}'] = student['inital']
        
        # Total price and price per student (only for first student)
        if i == 0:
            sheet[f'G{start_row}'] = student['electric_meter_before']
            sheet[f'H{start_row}'] = student['electric_meter_after']
            sheet[f'I{start_row}'] = student['used_unit']
        
        # Total price and price per student for each student
        sheet[f'J{row}'] = student['total_price']
        sheet[f'K{row}'] = student['price_divide_student']

def process_rooms(workbook, rooms, template_data):
    """Process rooms and create new sections in the same sheet"""
    rooms_per_section = 4
    total_sections = (len(rooms) + rooms_per_section - 1) // rooms_per_section
    sheet = workbook.active

    # Clear existing content except first section
    for row in range(26, sheet.max_row + 1):
        for col in range(1, 21):  # A to T
            cell = sheet[f"{get_column_letter(col)}{row}"]
            cell.value = None
            cell.border = None
            cell.fill = None
            cell.font = None
            cell.alignment = None

    # Process each section
    for section_num in range(total_sections):
        # Load fresh template for each section
        template_workbook = load_workbook('filestemplate/f2.xlsx')
        template_sheet = template_workbook.active
        
        start_row = section_num * 24 + 1
        setup_sheet_template(sheet, template_sheet, start_row)
        start_index = section_num * rooms_per_section
        
        # Add header title in A1 for each page
        header_cell = sheet[f'A{start_row}']
        header_cell.value = template_data['base_info']['header_title']
        header_cell.font = template_data['common']['font']
        
        # Fill room names and student data
        for i in range(rooms_per_section):
            if start_index + i < len(rooms):
                room = rooms[start_index + i]
                # Calculate the starting row for this room group
                room_start_row = start_row + (i * 4) + 4  # +4 because room names start at A5, A9, etc.
                fill_student_data(sheet, room, room_start_row)
        
        fill_room_names(sheet, rooms, start_index, start_row)
        
        # Close template workbook
        template_workbook.close()

@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    try:
        # Get data from request
        data = request.json
        
        # Validate required fields
        required_fields = ['building_name', 'date', 'unit_price', 'delay_price', 
                         'electric_detail', 'electric_detail_delay', 
                         'all_electric_total', 'all_delay_total', 'all_total', 'data_list']
        
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        # Format the data list
        formatted_data = format_data_list(data['data_list'])

        # Create template data
        template_data = {
            'common': {
                'font': Font(name='TH SarabunPSK', size=14),
                'font10': Font(name='TH SarabunPSK', size=10),
            },
            'base_info': {
                'building_name': data['building_name'],
                'date': data['date'],
                'unit_price': str(data['unit_price']),
                'delay_price': str(data['delay_price']),
                'electric_detail': data['electric_detail'],
                'electric_detail_delay': data['electric_detail_delay'],
                'all_electric_total': data['all_electric_total'],
                'all_delay_total': data['all_delay_total'],
                'all_total': data['all_total'],
            },
            'data_list': formatted_data
        }

        # Load workbooks
        workbook = load_workbook('filestemplate/f1.xlsx')
        bottom_workbook = load_workbook('filestemplate/f1-buttom.xlsx')
        
        sheet = workbook.active
        bottom_sheet = bottom_workbook.active

        # Copy print settings
        template_sheet = workbook.active
        sheet.page_setup = template_sheet.page_setup
        sheet.print_options = template_sheet.print_options
        sheet.print_title_rows = template_sheet.print_title_rows
        sheet.print_title_cols = template_sheet.print_title_cols
        sheet.print_area = template_sheet.print_area

        # Define styles
        font = template_data['common']['font']
        font10 = template_data['common']['font10']
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set bottom sheet values
        bottom_sheet['M1'].font = font10
        bottom_sheet['M1'].value = template_data['base_info']['all_electric_total']
        
        bottom_sheet['N1'].font = font10
        bottom_sheet['N1'].value = template_data['base_info']['all_delay_total']
        
        bottom_sheet['O1'].font = font10
        bottom_sheet['O1'].value = template_data['base_info']['all_total']
        
        bottom_sheet['I3'].font = font10
        bottom_sheet['I3'].value = f"=BAHTTEXT({template_data['base_info']['all_total']})"
        
        # Set other bottom sheet values
        bottom_sheet['B10'].font = font10
        bottom_sheet['B10'].value = template_data['base_info']['electric_detail']
        
        bottom_sheet['B11'].font = font10
        bottom_sheet['B11'].value = template_data['base_info']['electric_detail_delay']
        
        bottom_sheet['F10'].font = font10
        bottom_sheet['F10'].value = f"{template_data['base_info']['unit_price']} บาท/ต่อยูนิต"
        
        bottom_sheet['F11'].font = font10
        bottom_sheet['F11'].value = f"{template_data['base_info']['delay_price']} บาท/ต่อเดือน/คน"
        
        bottom_sheet['H10'].font = font10
        bottom_sheet['H10'].value = f"{template_data['base_info']['all_electric_total']} บาท"
        
        bottom_sheet['H11'].font = font10
        bottom_sheet['H11'].value = f"{template_data['base_info']['all_delay_total']} บาท"
        
        bottom_sheet['H12'].font = font10
        bottom_sheet['H12'].value = f"{template_data['base_info']['all_total']} บาท"

        # Set main sheet values
        sheet['A3'].font = font
        sheet['A3'].value = template_data['base_info']['building_name']
        
        sheet['J3'].font = font
        sheet['J3'].value = template_data['base_info']['date']
        
        sheet['I7'].font = font10
        sheet['I7'].value = f"@{template_data['base_info']['unit_price']}"
        
        sheet['N7'].font = font10
        sheet['N7'].value = f"@{template_data['base_info']['delay_price']}"

        # Fill data rows
        start_data_row = 7
        for row_num, row_data in template_data['data_list']['row'].items():
            current_row = int(row_num) + start_data_row
            
            for col in range(1, 18):
                col_letter = get_column_letter(col)
                cell = sheet[f"{col_letter}{current_row}"]
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for column, value in row_data['column'].items():
                cell = sheet[column + str(current_row)]
                cell.value = value
                cell.font = template_data['common']['font10']

        # Copy bottom sheet
        last_data_row = start_data_row + len(template_data['data_list']['row'])
        bottom_sheet_start_row = last_data_row + 1

        for row in range(1, 21):
            for col in range(1, 18):
                source_cell = bottom_sheet.cell(row=row, column=col)
                target_cell = sheet.cell(row=row + bottom_sheet_start_row - 1, column=col)
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)

        # Copy merged cells
        for merged_range in bottom_sheet.merged_cells.ranges:
            if merged_range.min_row <= 20 and merged_range.max_row <= 20:
                new_range = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row + bottom_sheet_start_row - 1}:{get_column_letter(merged_range.max_col)}{merged_range.max_row + bottom_sheet_start_row - 1}"
                sheet.merge_cells(new_range)

        # Copy dimensions
        for col in range(1, 18):
            col_letter = get_column_letter(col)
            if col_letter in bottom_sheet.column_dimensions:
                sheet.column_dimensions[col_letter].width = bottom_sheet.column_dimensions[col_letter].width

        for row in range(1, 21):
            if row in bottom_sheet.row_dimensions:
                sheet.row_dimensions[row + bottom_sheet_start_row - 1].height = bottom_sheet.row_dimensions[row].height

        # Save to BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='electric_bill.xlsx'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_image():
    # Check if an image file is included in the request
    if 'image' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files['image']
    
    # Validate that the file is an image
    if file and file.content_type.startswith('image'):
        try:
            # Open and resize the image if necessary
            image = Image.open(file)
            max_height = 640
            if image.height > max_height:
                height_percent = max_height / float(image.height)
                new_width = int(float(image.width) * height_percent)
                image = image.resize((new_width, max_height), Image.LANCZOS)
            
            # Convert image to numpy array for YOLO processing
            img_np = np.array(image.convert("RGB"))
            
            # Run YOLO prediction with confidence threshold
            results = model.predict(img_np, conf=0.5)
            
            # Prepare a blank background for annotated output
            bg_width, bg_height = 800, 600
            background = np.zeros((bg_height, bg_width, 3), dtype=np.uint8)
            
            # Draw bounding boxes and labels on the background image
            for result in results:
                boxes = result.boxes
                for box in boxes:
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    label = result.names[int(box.cls)]
                    
                    # Only label non-panel detections
                    if label != "panel":
                        label_position = (x1, y1 - 10 if y1 - 10 > 10 else y1 + 10)
                        cv2.putText(background, label, label_position, cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            
            # Save the annotated image temporarily
            labeled_img_path = "./src/tem/labeled_image.jpg"
            cv2.imwrite(labeled_img_path, background)
            
            # OCR processing on the labeled image
            ocr_image = Image.open(labeled_img_path)
            ocr_result = reader.readtext(np.array(ocr_image), detail=0)
            
            # Extract only numeric characters
            numbers_only = " ".join(re.findall(r'\d+', " ".join(ocr_result)))
            final = numbers_only.replace(" ", "")
            
            # crop only first 4 numbers
            final = final[:4]
            
            # Return OCR result as JSON
            return jsonify({"ocr_result": final}), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    else:
        return jsonify({"error": "File is not an image"}), 400

@app.route('/upload', methods=['GET'])
def hello_world():
    return jsonify("hello world /upload"), 200

@app.route('/generate-f2-excel', methods=['POST'])
def generate_f2_excel():
    try:
        # Get data from request
        data = request.json
        
        # Validate required fields
        required_fields = ['header_title', 'data_list']
        
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        # Create template data
        template_data = {
            'common': {
                'font': Font(name='TH SarabunPSK', size=14),
                'font10': Font(name='TH SarabunPSK', size=10),
            },
            'base_info': {
                'header_title': data['header_title'],
            },
            'data_list': {
                'room': data['data_list']
            }
        }

        # Load the workbook
        try:
            workbook = load_workbook('filestemplate/f2.xlsx')
        except FileNotFoundError as e:
            return jsonify({"error": f"Template file not found: {str(e)}"}), 500

        # Process the rooms from template_data
        rooms = template_data['data_list']['room']
        process_rooms(workbook, rooms, template_data)

        # Save to BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='electric_bill_f2.xlsx'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/generate-pay-report', methods=['POST'])
def generate_pay_report():
    try:
        # Get data from request
        data = request.json
        
        # Validate required fields
        required_fields = ['name-surename', 'student_id', 'major', 'faculty', 
                         'date_now', 'personal_id', 'room_number', 
                         'dormitory_name', 'electric_date_name', 'price']
        
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing required field: {field}"}), 400

        # Create result directory if it doesn't exist
        os.makedirs('result', exist_ok=True)

        template_path = os.path.abspath('filestemplate/pay_report.xlsx')
        output_path = os.path.abspath('result/pay_report_result.xlsx')

        try:
            # Copy template file
            shutil.copy2(template_path, output_path)
            
            # Load workbook using openpyxl
            wb = load_workbook(output_path)
            sheet = wb.active
            
            # Fill in the data
            sheet['B4'] = data['name-surename']
            sheet['B5'] = data['student_id']
            sheet['B6'] = data['major']
            sheet['B7'] = data['faculty']
            sheet['F4'] = data['date_now']
            sheet['F5'] = data['personal_id']
            sheet['F6'] = data['room_number']
            sheet['F7'] = data['dormitory_name']
            sheet['C10'] = data['electric_date_name']
            sheet['F10'] = data['price']

            # Add images using openpyxl
            # Resize images using PIL first
            pil_img1 = Image.open('filestemplate/photo1.png')
            pil_img2 = Image.open('filestemplate/photo2.png')
            pil_img3 = Image.open('filestemplate/photo3.png')
            
            # Resize images
            pil_img1 = pil_img1.resize((100, 75), Image.Resampling.LANCZOS)
            pil_img2 = pil_img2.resize((150, 100), Image.Resampling.LANCZOS)
            pil_img3 = pil_img3.resize((150, 100), Image.Resampling.LANCZOS)
            # Save resized images temporarily
            temp_img1_path = 'result/temp_img1.png'
            temp_img2_path = 'result/temp_img2.png'
            temp_img3_path = 'result/temp_img3.png'
            pil_img1.save(temp_img1_path)
            pil_img2.save(temp_img2_path)
            pil_img3.save(temp_img3_path)
            
            # Load resized images into Excel
            img1 = XLImage(temp_img1_path)
            img2 = XLImage(temp_img2_path)
            img3 = XLImage(temp_img3_path)
            # Add photo1.png to A27
            sheet.add_image(img1, 'A27')

            # Add photo2.png to A1 and A18
            sheet.add_image(img2, 'A1')
            sheet.add_image(img3, 'A19')

            # Get QR code data from B33
            qr_data = sheet['B33'].value
            if not qr_data:
                qr_data = "No data"  # Default value if B33 is empty

            # Generate QR Code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(str(qr_data))  # Convert to string to ensure compatibility
            qr.make(fit=True)

            # Create QR code image
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            # Resize QR code
            qr_img = qr_img.resize((100, 100), Image.Resampling.LANCZOS)
            
            # Save QR code temporarily
            qr_path = os.path.abspath('result/qr_temp.png')
            qr_img.save(qr_path)

            # Add QR code to Excel using openpyxl
            qr_excel_img = XLImage(qr_path)
            sheet.add_image(qr_excel_img, 'A32')
            
            # Save the workbook
            wb.save(output_path)
            
            # Read the generated file
            with open(output_path, 'rb') as f:
                excel_file = BytesIO(f.read())
            
            # Clean up temporary files
            os.remove(output_path)
            os.remove(qr_path)
            os.remove(temp_img1_path)
            os.remove(temp_img2_path)
            os.remove(temp_img3_path)
            
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='pay_report.xlsx'
            )

        except FileNotFoundError:
            return jsonify({"error": f"Template file not found at {template_path}"}), 404
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=9002)